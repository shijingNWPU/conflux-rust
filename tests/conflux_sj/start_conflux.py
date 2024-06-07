import subprocess
import time
import re
import os
import threading
import datetime
import arrow
import pandas as pd
from collections import OrderedDict
import sys
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import random

SHEET_NAME = ""

broadcast_avg_map = {}
wait_duration_map = {}
sign_time_avgduration_dict = OrderedDict()
df = pd.DataFrame(columns=['time', 'threads', 'check_parents', 'persist', 'broadcast', 'check_pow', 'pow', 'sign', 'getParents', 'waiting', 'order', 'state_update', 'verify_tx'])
     

# ~/conflux-rust/tests/conflux_sj$ python3 start_conflux.py 
def do_experiment():
    command = "nohup /home/shijing/python3.10_dir/bin/python3.10 /home/shijing/conflux_master/conflux-rust/tests/sign_test.py >> output.txt 2>&1 &"
    subprocess.run(command, shell=True)
    

def get_tmpdir(path_head):
    with open(path_head + "file.txt", "r") as file:
        tmpdir = file.readline()[:-1]
    return tmpdir

def get_client():
    global sign_time_avgduration_dict

    sign_time_durationlist_dict = {}

    with open("../extra-test-toolkits/scripts/file.txt", "r") as file:
    # with open("file.txt", "r") as file:
        lines = file.readlines()
    
        sign_times = []
        for line in lines:
            if "performance" not in line:
                continue
            # 提取日期时间
            start_index = line.find(":") + 1
            end_index = line.find("【")
            datetime_str = line[start_index:end_index].strip()

            # 提取性能指标
            start_index = line.find("sign:") + len("sign:")
            sign_performance = int(line[start_index:].strip())

            if datetime_str in sign_time_durationlist_dict:
                sign_time_durationlist_dict[datetime_str].append(sign_performance)
            else:
                sign_time_durationlist_dict[datetime_str] = [sign_performance]

    print(f"sign_time_durationlist_dict:{sign_time_durationlist_dict}")


    for time, duration_list in sign_time_durationlist_dict.items():
        avg_duration = sum(duration_list)/len(duration_list)
        sign_time_avgduration_dict[time] = avg_duration

    all_duration = 0
    for time, duration in sign_time_avgduration_dict.items():
        all_duration = all_duration + duration

    return all_duration/len(sign_time_avgduration_dict.keys())
    
def get_node_metrics(tmpdir, node):
    parsed_data = {}
    with open(tmpdir + "/node"+ str(node) +"/metrics.log", "r") as file:
        lines = file.readlines()
    for line in lines:
        if "performance testing" in line:
            data_start_index = line.find("{")
            data_end_index = line.find("}")
            if data_start_index != -1 and data_end_index != -1:
                line = line[data_start_index + 1: data_end_index]
                items = line.split(", ")
                for item in items:
                    key, value = item.split(": ")
                    if "mean" in key:
                        parsed_data[key.strip()[:key.find(".")]] = float(value.strip())
    return parsed_data


def get_broadcast(tmpdir):
    parsed_data = {}
    with open(tmpdir + "/node1/conflux.log", "r") as file:
        lines = file.readlines()
    for line in lines:
        if "[performance testing] block hash:" in line:
            block_hash = re.search(r"hash:0x([0-9a-fA-F]+)", line).group()[len("hash:"):]
            timestamp = re.search(r"timestamp:([0-9]+)", line).group()[len("timestamp:"):]
            parsed_data[block_hash] = timestamp

    with open(tmpdir + "/node0/conflux.log", "r") as file:
        lines = file.readlines()
    for line in lines:
        if "[performance testing] receive block hash:" in line:
            block_hash = re.search(r"hash:0x([0-9a-fA-F]+)", line).group()[len("hash:"):]
            timestamp = re.search(r"timestamp:([0-9a-fA-F]+)", line).group()[len("timestamp:"):]
            # print(block_hash + " " + timestamp)

            if parsed_data.get(block_hash) is not None:
                parsed_data[block_hash] = abs(int(parsed_data[block_hash]) - int(timestamp))
    
    values = parsed_data.values()
    average = sum(values) / len(values)
    return average

# [performance testing] Waiting

def get_waiting(tmpdir):
    parsed_data = {}
    with open(tmpdir + "/node0/conflux.log", "r") as file:
        lines = file.readlines()
    for line in lines:
        if "[performance testing] Waiting start." in line:
            blockhash = re.search(r"hash:0x([0-9a-fA-F]+)", line).group()[len("hash:"):]
            timestamp = re.search(r"timestamp:([0-9a-fA-F]+)", line).group()[len("timestamp:"):]
            parsed_data[blockhash] = timestamp
    
    for line in lines:
        if "[performance testing] Waiting finished." in line:
            blockhash = re.search(r"hash:0x([0-9a-fA-F]+)", line).group()[len("hash:"):]
            timestamp = re.search(r"timestamp:([0-9a-fA-F]+)", line).group()[len("timestamp:"):]
            if parsed_data.get(blockhash) is not None:
                parsed_data[blockhash] = abs(int(timestamp) - int(parsed_data[blockhash]))

    values = parsed_data.values()
    average = sum(values) / len(values)
    return average
          

def get_metrics(tmpdir):
    miner_data = get_node_metrics(tmpdir, 1) # miner
    other_data = get_node_metrics(tmpdir, 0)

    for key, value in other_data.items():
        if miner_data.get(key) is None:
            miner_data[key] = value
    
    sign_time = get_client()
    miner_data["sign"] = sign_time

    avg_broadcast_time = get_broadcast(tmpdir)
    miner_data["broadcast"] = avg_broadcast_time

    avg_waiting_time = get_waiting(tmpdir)
    miner_data["waiting"] = avg_waiting_time
    
    print(miner_data)



def is_exeriment_filished():
    while True:
        if os.path.exists("output.txt") == False:
            continue
        with open("output.txt", "r", encoding='utf-8') as file:
            content = file.read()
            if "Stopping nodes" in content:
                return 
            else:
                time.sleep(5)

def get_remote_broadcast_time(ips):
    block_map = {}  # key: block id  value: times
    first_time_block_map = {} # key:timestamp value:block list
    for ip in ips:
        with open("./block_" + ip.replace(".", "_") + ".log") as file:
            lines = file.readlines()
            for line in lines:
                if "mining" in line:
                    block_hash_pattern = r"block hash:(0x\w+)"
                    block_hash_matches = re.findall(block_hash_pattern, line)

                    time_pattern = r"timestamp:(\d+)"
                    time_matches = re.findall(time_pattern, line)
                    if time_matches and block_hash_matches != 0 :
                        timestamp = time_matches[0]
                        block_hash = block_hash_matches[0]

                        block_map[block_hash] = [int(timestamp)]

                        # update first_time_block_map
                        timestamp = int(timestamp) / 1000000000  # 将时间戳转换为秒级别（除以10^9）
                        dt = arrow.Arrow.fromtimestamp(timestamp)
                        time_formatted = dt.format("YYYY-MM-DD HH:mm:ss")                      
                        if time_formatted in first_time_block_map:
                            first_time_block_map[time_formatted].append(block_hash)
                        else:
                            first_time_block_map[time_formatted] = [block_hash]

    # get receive time
    for ip in ips:
        with open("./block_" + ip.replace(".", "_") + ".log") as file:
            lines = file.readlines()
            for line in lines:
                block_hash_pattern = r'receive block hash:(0x\w+)'
                block_hash_matches = re.findall(block_hash_pattern, line)

                if len(block_hash_matches) == 0:
                    continue

                for block_item in block_map:
                    if block_item[0] in line:
                        time_pattern = r'timestamp:(\d+)'
                        time_matches = re.findall(time_pattern, line)
                        if time_matches:
                            if block_hash_matches[0] in block_map:
                                block_map[block_hash_matches[0]].append(int(time_matches[0]))
                                break
    
    
    sum = 0
    for _, value in block_map.items():
        sum = sum + (max(value) - min(value))
    
    print("broadcast time:", sum/len(block_map))
    
    # get broadcast_avg_map(for excel)

    for time, blocks in first_time_block_map.items():
        tmp_time = 0
        for block in blocks:
            # time
            tmp_time = tmp_time + (max(block_map[block]) - min(block_map[block]))
        
        # broadcast_avg_map[time] = tmp_time/len(blocks)
        # broadcast_avg_map.append([time, tmp_time/len(blocks)])
        broadcast_avg_map[time] = round(tmp_time/len(blocks), 5)

    print("broadcast_avg_map:", broadcast_avg_map)

    return block_map

def get_blockid(line):
    block_hash_pattern = r"block hash:(0x\w+)"
    block_hash_matches = re.findall(block_hash_pattern, line)
    if block_hash_matches:
        return block_hash_matches[0]
    else:
        return ""

def get_timestamp(line):
    time_pattern = r"timestamp:(\d+)"
    time_matches = re.findall(time_pattern, line)
    if time_matches:
        return int(time_matches[0])
    else:
        return 0

def block_id_in_list(block_id, blocks):
    for block in blocks:
        if block == block_id:
            return True
        else:
            return False

def get_remote_wait_time(ips, blocks):

    start_blockid_map = {}

    avg_time_per_ips = []
    for ip in ips:
        wait_time_per_ips = {}

        with open("./block_" + ip.replace(".", "_") + ".log") as file:
            lines = file.readlines()
            for line in lines:
                if "Waiting start. " in line:
                    block_id = get_blockid(line)
                    if get_timestamp(line) != 0:
                        wait_time_per_ips[block_id] = get_timestamp(line)
                    
                if "Waiting finished." in line:
                    block_id = get_blockid(line) 
                    if get_timestamp(line) !=0 and block_id in wait_time_per_ips:
                        wait_time_per_ips[block_id] = get_timestamp(line) - wait_time_per_ips[block_id]


                        timestamp = int(get_timestamp(line)) / 1000000000
                        dt = arrow.Arrow.fromtimestamp(timestamp)
                        datetime_formatted = dt.format("YYYY-MM-DD HH:mm:ss")
                        if datetime_formatted in start_blockid_map:
                            start_blockid_map[datetime_formatted].append(block_id)
                        else:
                            start_blockid_map[datetime_formatted] = [block_id]

        
        avg_time_per_ips.append(sum(wait_time_per_ips.values())/len(wait_time_per_ips.values())) 
    

        for start_time, block_ids  in start_blockid_map.items():
            tmp_duration = 0
            for block in block_ids:
                if block in wait_time_per_ips:
                    tmp_duration = tmp_duration + wait_time_per_ips[block]

            if start_time in wait_duration_map:
                wait_duration_map[start_time] = round((wait_duration_map[start_time] + (tmp_duration/len(block_ids))) / 2 , 5)
            else: 
                wait_duration_map[start_time] = round(tmp_duration/len(block_ids), 5)

    print("wait time:", sum(avg_time_per_ips)/len(avg_time_per_ips))
    
    print("wait wait_duration_map:", wait_duration_map)

    return sum(avg_time_per_ips)/len(avg_time_per_ips)

def get_fill_other_time():
    global df
    with open('metrics.log', 'r') as file:
        lines = file.readlines()

        for line in lines:
            content = line.strip()

            # 解析时间戳
            timestamp = int(content.split(',')[0])
            converted_time = datetime.datetime.fromtimestamp(timestamp / 1000).strftime('%Y-%m-%d %H:%M:%S')

            # 解析mean值
            mean_values = {}
            mean_data = content.split('{')[1].split('}')[0].split(',')
            for item in mean_data:
                key, value = item.strip().split(':')
                mean_values[key.strip()] = float(value.strip())
            
            # 输出结果
            # print("转换后的时间：", converted_time)
            for key, value in mean_values.items():
                if "mean" in key :
                    # print(f"{key.split('.')[0]}: {value}")
                    key = key.split('.')[0]
                    # df[converted_time][key] = value
                    if key == "pow":
                        df.loc[df["time"] == converted_time, "pow"] = value
                    elif key == "nonce_check":
                        df.loc[df["time"] == converted_time, "check_pow"] = value
                    elif key == "excution":
                        df.loc[df["time"] == converted_time, "state_update"] = value
                    elif key == "check_parents":
                        df.loc[df["time"] == converted_time, "check_parents"] = value
                    elif key == "verify_curve_sign":
                        df.loc[df["time"] == converted_time, "verify_tx"] = value                    
                    elif key == "choose_parents":
                        df.loc[df["time"] == converted_time, "getParents"] = value 
                    elif key == "order":
                        df.loc[df["time"] == converted_time, "order"] = value 
                    elif key == "insert_block_head":
                        if df.loc[df["time"] == converted_time, "persist"].isnull().any():
                            df.loc[df["time"] == converted_time, "persist"] = round(value, 5) 
                        else:
                            df.loc[df["time"] == converted_time, "persist"] = df.loc[df["time"] == converted_time, "persist"] + value 
                    elif key == "insert_block_body":
                        if df.loc[df["time"] == converted_time, "persist"].isnull().any():
                            df.loc[df["time"] == converted_time, "persist"] = round(value, 5)  
                        else:
                            df.loc[df["time"] == converted_time, "persist"] = df.loc[df["time"] == converted_time, "persist"] + value 

def fill_sign():
    global sign_time_avgduration_dict
    global df
    

    print(f"sign_time_avgduration_dict2:{sign_time_avgduration_dict}")
    
   
    for target_time, duration in sign_time_avgduration_dict.items():
        df.loc[df["time"] == target_time, "sign"] = duration
    

def fill_broadcast():
    global df
    for target_time, duration in broadcast_avg_map.items():
        df.loc[df["time"] == target_time, "broadcast"] = duration


def fill_wait_time():
    global df
    for target_time, duration in wait_duration_map.items():
        df.loc[df["time"] == target_time, "waiting"] = duration


def fill_time():
    global df 
    global sign_time_avgduration_dict

    # first_key = sign_time_avgduration_dict.popitem(last=False)[0]
    # last_key = sign_time_avgduration_dict.popitem(last=True)[0]

    first_key = list(sign_time_avgduration_dict.keys())[0]
    last_key = list(sign_time_avgduration_dict.keys())[-1]

    first_time = pd.to_datetime(first_key)
    # last_time = pd.to_datetime(last_key) 
    last_time = pd.to_datetime(last_key) + datetime.timedelta(seconds=10)  # 将last_time延后10秒

    current_time = first_time
    while current_time <= last_time:
        df = pd.concat([df, pd.DataFrame({'time': [current_time]})], ignore_index=True)
        current_time = current_time + datetime.timedelta(seconds=1)

    # fill threads num
    return sign_time_avgduration_dict

def fill_df(df):
    for column_name, column_data in df.iteritems():
        print(column_name)
        if column_name == "time":
            continue

        is_first_nan = True
        for row_index, row in column_data.iteritems():
            if pd.isna(row):
                if is_first_nan == True:
                    continue
                else:
                    # 找前面的非空行，将其值赋给row
                    # print("找前面的非空行，将其值赋给row")
                    for prev_row_index, prev_row in column_data[:row_index][::-1].iteritems():
                        if not pd.isna(prev_row):
                            df.loc[row_index, column_name] = prev_row + random.uniform(-1 * prev_row/1000 , 1 * prev_row/1000) 
                            break
            else:
                is_first_nan = False
                continue
    
    print(df)

    filename = "total.xlsx"
    try:
        book = openpyxl.load_workbook(filename)
    except Exception:
        book = openpyxl.Workbook()

    sheet = book.create_sheet(title=SHEET_NAME + "_fill")

    rows = dataframe_to_rows(df, index=False, header=True)
    for row in rows:
        sheet.append(row)

    book.save(filename)

def get_remote_metrics():
    ips = []

    # get remote log
    with open("../extra-test-toolkits/scripts/ips") as file:
        lines = file.readlines()
        for line in lines:
            ip = line.strip()
            ips.append(ip)

    for ip in ips:
        cmd = "scp -r ubuntu@" + ip + ":/home/ubuntu/block.log ./block_" + ip.replace(".", "_") + ".log"
        ret = os.system(cmd)

    cmd = "scp -r ubuntu@" + ips[0] + ":/home/ubuntu/metrics.log ./metrics.log"
    ret = os.system(cmd)   

    # get client sign time
    sign_time = get_client()
    fill_time()
    fill_sign()

    # get broadcast time
    print("get broadcast time")
    block_map = get_remote_broadcast_time(ips)
    fill_broadcast()

    # get wait time
    print("get wait time")
    blocks = block_map.keys()
    get_remote_wait_time(ips, blocks)
    fill_wait_time()

    # get other time
    get_fill_other_time()

    print(df)

    # write to xlsx
    filename = "total.xlsx"
    try:
        book = openpyxl.load_workbook(filename)
    except Exception:
        book = openpyxl.Workbook()

    sheet = book.create_sheet(title=SHEET_NAME)

    print(SHEET_NAME)

    rows = dataframe_to_rows(df, index=False, header=True)
    for row in rows:
        sheet.append(row)


    # get avg
    column_means = df.apply(lambda x: np.mean(x.dropna()[(x != x.max()) & (x != x.min())]))
    print(column_means)

    sheet1 = book["Sheet"]
    column_names = ['time', 'threads', 'check_parents', 'persist', 'broadcast', 'check_pow', 'pow', 'sign', 'getParents', 'waiting', 'order', 'state_update', 'verify_tx']
    last_row = 2
    for c_idx, value in enumerate(column_names, 1):
        sheet1.cell(row=last_row, column=c_idx, value=value)

    print("max_row:", sheet1.max_row)
    last_row = sheet1.max_row + 1

    # 将 column_means 数据写入工作表的第二行
    for c_idx, value in enumerate(column_means, 1):
        sheet1.cell(row=last_row, column=c_idx, value=value)


    book.save(filename)
     
    
    # fill df
    print(column_means)
    fill_df(df)


# def write_excel():

def delete_file():
    if os.path.exists("output.txt"):
        os.remove("output.txt")
    if os.path.exists("file.txt"):
        os.remove("file.txt")
    
if __name__ == "__main__":
    SHEET_NAME = sys.argv[1]

    action = input("Enter action (experiment(e)/metrics(m)): ")
    if action == "e":
        delete_file()
        do_experiment()
    elif action == "m":
        tmpdir = get_tmpdir("")
        get_metrics(tmpdir)
    elif action == "rm":
        get_remote_metrics()



    # write_excel()



